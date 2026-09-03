# -*- coding: utf-8 -*-
"""成果物を1つの xlsx にまとめる。

CSVが4つに分かれていて貼り付け先を間違えやすかったので、
「検証結果」「G列訂正」「要確認」「先方確認事項」を1ブックに集約する。
集計シートは COUNTIF で検証結果シートを参照しているので、
判定を手で直せば数字が追随する。
"""
import csv, json, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

R = "tokium/results/"
OUT = R + "TOKIUM_直アポ大_検証結果.xlsx"
FONT = "Arial"

NAVY = "1F3864"; GREY = "F2F2F2"; RED = "C00000"; AMBER = "FFF2CC"
head_fill = PatternFill("solid", fgColor=NAVY)
head_font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
base = Font(name=FONT, size=10)
thin = Side(style="thin", color="BFBFBF")
box = Border(left=thin, right=thin, top=thin, bottom=thin)


def read(path):
    return list(csv.DictReader(open(R + path, encoding="utf-8-sig")))


def sheet(wb, title, cols, rows, widths, light=False):
    """ヘッダ付きの表を1枚作る。cols は (見出し, dictキー) の並び。

    light=True は罫線と折り返しを省く。4,422行に個別書式を付けると
    LibreOffice の再計算が3分で終わらないため、大きいシートで使う。
    """
    ws = wb.create_sheet(title)
    for c, (label, _) in enumerate(cols, 1):
        cell = ws.cell(1, c, label)
        cell.fill, cell.font, cell.border = head_fill, head_font, box
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    top = Alignment(vertical="top")
    wrap = Alignment(vertical="top", wrap_text=True)
    for r, row in enumerate(rows, 2):
        for c, (_, key) in enumerate(cols, 1):
            v = row.get(key, "")
            if key == "行" and str(v).strip():
                v = int(v)
            cell = ws.cell(r, c, v)
            cell.font = base
            if light:
                cell.alignment = top
            else:
                cell.border = box
                cell.alignment = wrap if c >= len(cols) - 1 else top
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rows)+1}"
    ws.row_dimensions[1].height = 28
    return ws


wb = Workbook()
wb.remove(wb.active)

# ── 1. 読み方 ────────────────────────────────────────────────
full = read("IJ_paste_FULL.csv")
ws = wb.create_sheet("読み方")
ws.column_dimensions["A"].width = 104
GUIDE = [
    ("TOKIUM 直アポ大 — G列・H列の検証結果", "title"),
    ("", ""),
    ("全4,422行を検証しました。判断基準は御社からご提供いただいた RULES.md が唯一の根拠です。", ""),
    ("", ""),
    ("シートの使い方", "h"),
    ("① 検証結果 … 全4,422行。I列（検証結果）とJ列（出典）をこのままシートに貼れます。", ""),
    ("      元のシートのI2セルに、H列とI列の2列ぶんを貼ってください。行は欠番なく1行目=行2です。", ""),
    ("② 集計 … 判定の内訳。COUNTIFで①を参照しているので、①を直せば数字が追随します。", ""),
    ("③ G列訂正 … G列の記載に誤りがあった行。売上高の数値・決算期・単体連結の別のいずれかが違います。", ""),
    ("④ 要確認（薄いマージン）… 架電可だが300億円の基準に近く、数値が少し違うと判定が反転する行。", ""),
    ("⑤ 先方確認事項 … 御社にご判断いただきたい点。特に §7-1 は577行の帰趨を左右します。", ""),
    ("⑥ 監査で落とした行 … 一度「架電可」と判定したが、再検証で誤りと判明し取り下げた行。", ""),
    ("", ""),
    ("判定ラベルの意味", "h"),
    ("架電可 … G/H列の記載が正しく、架電してよい", ""),
    ("架電可（G列訂正あり）… 架電してよいが、G列の数値等に訂正が必要（③参照）", ""),
    ("架電可（回収）… 元は△×で除外されていたが、承認ソースを見つけて架電可に戻した", ""),
    ("架電不可（△要再確認）… 承認ソースで売上高を確認できない。§3-5によりリストに載せられない", ""),
    ("架電不可（×要修正）… 300億円未満・法人格NG等が確定した", ""),
    ("架電不可（元×判定が妥当と確認）… 元から除外されており、その除外が正しいと確認した", ""),
    ("", ""),
    ("この検証の限界", "h"),
    ("作業環境のネットワーク制限により、PDFやWebページの本文を直接開けませんでした。", ""),
    ("検証はすべて検索結果のスニペット経由です。「承認ソースが存在することは確認できたが", ""),
    ("本文の数値に到達できない」という理由で△に留まった行が相当数あります（⑤に記載）。", ""),
    ("本文を開ける環境であれば、これらの多くは回収または確定ができます。", ""),
]
for i, (text, kind) in enumerate(GUIDE, 1):
    c = ws.cell(i, 1, text)
    if kind == "title":
        c.font = Font(name=FONT, bold=True, size=15, color=NAVY)
    elif kind == "h":
        c.font = Font(name=FONT, bold=True, size=11, color=NAVY)
    else:
        c.font = base
    c.alignment = Alignment(wrap_text=True, vertical="top")

# ── 2. 検証結果（全行）────────────────────────────────────────
cols = [("行", "行"), ("取引先ID", "取引先ID"), ("法人番号", "法人番号"),
        ("取引先名", "取引先名"), ("元H判定", "元H判定"), ("最終判定", "最終判定"),
        ("架電可", "架電可"), ("I列_検証結果", "I列_検証結果"), ("J列_ソース", "J列_ソース")]
ws = sheet(wb, "検証結果", cols, full, [7, 12, 15, 30, 8, 22, 7, 95, 55], light=True)
ng = Font(name=FONT, size=10, color=RED)
for r in range(2, len(full) + 2):
    if ws.cell(r, 7).value != "○":
        ws.cell(r, 6).font = ng

# ── 3. 集計（検証結果シートを参照）─────────────────────────────
ws = wb.create_sheet("集計")
n = len(full) + 1
ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 12
ws.column_dimensions["C"].width = 10
ws.column_dimensions["D"].width = 62
for c, label in enumerate(["区分", "行数", "割合", "意味"], 1):
    cell = ws.cell(1, c, label)
    cell.fill, cell.font, cell.border = head_fill, head_font, box
LABELS = [
    ("架電可", "G/H列の記載が正しく、そのまま架電してよい"),
    ("架電可（G列訂正あり）", "架電してよいが、G列に訂正が必要"),
    ("架電可（回収）", "元は除外されていたが、承認ソースを見つけて戻した"),
    ("架電不可（△要再確認）", "承認ソースで売上高を確認できない。§3-5により載せられない"),
    ("架電不可（元×判定が妥当と確認）", "元から除外されており、その除外が正しいと確認した"),
    ("架電不可（×要修正）", "300億円未満・法人格NG等が確定した"),
    ("架電不可（元△判定が妥当と確認）", "元の確認不可判定が正しいと確認した"),
    ("架電不可（保留）", "上記のいずれにも当てはまらない"),
]
# 割合の分母は必ず「全行」セル。相対で書くと行ごとに分母がずれるので絶対参照で固定する。
OK_ROW = 2 + len(LABELS)          # 架電可 合計
TOTAL_ROW = OK_ROW + 1            # 全行
NG_ROW = TOTAL_ROW + 1            # 架電不可 合計
DEN = f"$B${TOTAL_ROW}"

r = 2
for label, meaning in LABELS:
    ws.cell(r, 1, label).font = base
    ws.cell(r, 2, f'=COUNTIF(検証結果!$F$2:$F${n},A{r})').font = base
    ws.cell(r, 3, f'=IFERROR(B{r}/{DEN},0)').font = base
    ws.cell(r, 3).number_format = "0.0%"
    ws.cell(r, 4, meaning).font = base
    for c in range(1, 5):
        ws.cell(r, c).border = box
        ws.cell(r, c).alignment = Alignment(vertical="top", wrap_text=(c == 4))
    r += 1
bold = Font(name=FONT, bold=True, size=10)
for r, (title, formula) in zip(
    (OK_ROW, TOTAL_ROW, NG_ROW),
    [("架電可 合計", f'=COUNTIF(検証結果!$G$2:$G${n},"○")'),
     ("全行", f'=COUNTA(検証結果!$A$2:$A${n})'),
     ("架電不可 合計", f'=B{TOTAL_ROW}-B{OK_ROW}')]):
    ws.cell(r, 1, title).font = bold
    ws.cell(r, 2, formula).font = bold
    ws.cell(r, 3, f'=IFERROR(B{r}/{DEN},0)').font = bold
    ws.cell(r, 3).number_format = "0.0%"
    for c in range(1, 5):
        ws.cell(r, c).fill = PatternFill("solid", fgColor=GREY)
        ws.cell(r, c).border = box

# ── 4〜5. G列訂正 / 要確認 ─────────────────────────────────────
g = read("G_column_corrections.csv")
sheet(wb, "G列訂正", [("行", "行"), ("取引先ID", "取引先ID"), ("法人番号", "法人番号"),
                     ("取引先名", "取引先名"), ("最終判定", "最終判定"),
                     ("G列の訂正指示", "G列の訂正指示"), ("I列_検証結果", "I列_検証結果")],
      g, [7, 12, 15, 30, 22, 52, 95])

t = read("thin_margin_watchlist.csv")
sheet(wb, "要確認（薄いマージン）",
      [("行", "行"), ("取引先名", "取引先名"), ("法人番号", "法人番号"),
       ("採用値_億円", "採用値_億円"), ("金額の出所", "金額の出所"),
       ("余裕率", "余裕率"), ("G列", "G列"), ("I列", "I列")],
      t, [7, 30, 15, 12, 26, 10, 30, 95])

# ── 6. 先方確認事項 ───────────────────────────────────────────
ws = wb.create_sheet("先方確認事項")
ws.column_dimensions["A"].width = 112
md = open(R + "OPEN_QUESTIONS.md", encoding="utf-8").read()
r = 1
for line in md.split("\n"):
    s = line.rstrip()
    if s.strip() in ("---", ""):
        r += 1
        continue
    m = re.match(r"^(#{1,3})\s*(.*)$", s)
    if m:
        lv, text = len(m.group(1)), m.group(2)
        c = ws.cell(r, 1, text)
        c.font = Font(name=FONT, bold=True, color=NAVY,
                      size={1: 14, 2: 12, 3: 11}[lv])
    else:
        text = re.sub(r"\*\*(.+?)\*\*", r"\1", s)
        text = re.sub(r"`(.+?)`", r"\1", text)
        c = ws.cell(r, 1, text)
        c.font = base
    c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

# ── 7. 監査で落とした行 ───────────────────────────────────────
ov = json.load(open("tokium/raw/audit_overrides.json", encoding="utf-8"))
by = {int(x["行"]): x for x in full}
aud = []
for k in sorted(ov, key=int):
    n_ = int(k)
    aud.append({"行": n_, "取引先名": by[n_]["取引先名"], "法人番号": by[n_]["法人番号"],
                "最終判定": by[n_]["最終判定"], "I列_検証結果": by[n_]["I列_検証結果"],
                "J列_ソース": by[n_]["J列_ソース"]})
ws = sheet(wb, "監査で落とした行",
           [("行", "行"), ("取引先名", "取引先名"), ("法人番号", "法人番号"),
            ("最終判定", "最終判定"), ("I列_検証結果", "I列_検証結果"), ("J列_ソース", "J列_ソース")],
           aud, [7, 30, 15, 22, 100, 50])
for r_ in range(2, len(aud) + 2):
    if "架電不可" in str(ws.cell(r_, 4).value):
        ws.cell(r_, 4).font = ng
        for c in range(1, 7):
            ws.cell(r_, c).fill = PatternFill("solid", fgColor=AMBER)

wb.save(OUT)

# ── 数式のキャッシュ値を書き込む ────────────────────────────────
# openpyxl は数式を文字列で書くだけでキャッシュ値を持たない。この環境の
# LibreOffice は3行のブックでも再計算が終わらず recalc.py が使えないため、
# 同じ集計をPythonで計算してXMLに <v> として埋める。
# 数式自体は残るので、判定を書き換えればExcel側で再計算される。
import zipfile, shutil, os
from collections import Counter

cnt = Counter(x["最終判定"] for x in full)
n_ok = sum(1 for x in full if x["架電可"] == "○")
n_all = len(full)
cache = {}
for i, (label, _) in enumerate(LABELS):
    r_ = 2 + i
    cache[f"B{r_}"] = cnt.get(label, 0)
    cache[f"C{r_}"] = cnt.get(label, 0) / n_all
cache[f"B{OK_ROW}"] = n_ok
cache[f"B{TOTAL_ROW}"] = n_all
cache[f"B{NG_ROW}"] = n_all - n_ok
cache[f"C{OK_ROW}"] = n_ok / n_all
cache[f"C{TOTAL_ROW}"] = 1.0
cache[f"C{NG_ROW}"] = (n_all - n_ok) / n_all

# 集計シートが何番目の sheet*.xml かを名前から引く
idx = wb.sheetnames.index("集計") + 1
target = f"xl/worksheets/sheet{idx}.xml"
tmp = OUT + ".tmp"
with zipfile.ZipFile(OUT) as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
    for item in zin.infolist():
        data = zin.read(item.filename)
        if item.filename == target:
            xml = data.decode("utf-8")
            for ref, val in cache.items():
                v = repr(round(val, 12)) if isinstance(val, float) else str(val)
                xml = re.sub(
                    r'(<c r="%s"[^>]*>)(<f>.*?</f>)(?!<v>)' % ref,
                    lambda m: m.group(1) + m.group(2) + "<v>%s</v>" % v,
                    xml, count=1)
            data = xml.encode("utf-8")
        zout.writestr(item, data)
os.replace(tmp, OUT)

# 埋め込んだ値を読み戻して、集計が本体と一致するか確かめる
from openpyxl import load_workbook
chk = load_workbook(OUT, data_only=True)["集計"]
assert chk.cell(TOTAL_ROW, 2).value == n_all, "全行が一致しない"
assert chk.cell(OK_ROW, 2).value == n_ok, "架電可が一致しない"
assert chk.cell(NG_ROW, 2).value == n_all - n_ok, "架電不可が一致しない"
assert sum(chk.cell(2 + i, 2).value for i in range(len(LABELS))) == n_all, "内訳の合計が全行に一致しない"
print("集計キャッシュ検証OK: 全行%d / 架電可%d / 架電不可%d" % (n_all, n_ok, n_all - n_ok))

print("->", OUT)
print("シート:", " / ".join(wb.sheetnames))
print("検証結果 %d行 / G列訂正 %d行 / 要確認 %d行 / 監査 %d行" % (len(full), len(g), len(t), len(aud)))
